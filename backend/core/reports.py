"""
ISOSmart Business Reports
Generates PDF, XLSX, and CSV reports for the three main report types:
  - sgq_executive : Estado SGQ Ejecutivo
  - risks         : Riesgos y Oportunidades
  - objectives    : Objetivos y Desempeño
"""
import csv
import io
import datetime

from django.utils import timezone

# ---------------------------------------------------------------------------
# Report type / format labels
# ---------------------------------------------------------------------------

REPORT_LABELS = {
    'sgq_executive': 'Estado SGQ Ejecutivo',
    'risks':         'Riesgos y Oportunidades',
    'objectives':    'Objetivos y Desempeño',
}

# ---------------------------------------------------------------------------
# Data helpers
# ---------------------------------------------------------------------------

def _build_date_filter(date_field, date_from, date_to):
    """Return a dict suitable for queryset.filter(**…)."""
    f = {}
    if date_from:
        f[f'{date_field}__date__gte'] = date_from
    if date_to:
        f[f'{date_field}__date__lte'] = date_to
    return f


def _risk_data(organization, date_from, date_to, status_filter=None):
    from .models import RiskMatrix
    qs = RiskMatrix.objects.filter(organization=organization)
    qs = qs.filter(**_build_date_filter('detection_date', date_from, date_to))
    if status_filter:
        qs = qs.filter(status=status_filter)
    return list(qs.order_by('-detection_date'))


def _objective_data(organization, date_from, date_to, status_filter=None):
    from .models import QualityObjective
    qs = QualityObjective.objects.filter(organization=organization)
    qs = qs.filter(**_build_date_filter('created_at', date_from, date_to))
    if status_filter:
        qs = qs.filter(status=status_filter)
    return list(qs.order_by('-created_at'))


def _executive_stats(organization, date_from, date_to):
    from .models import RiskMatrix, QualityObjective, Document, StakeholderProfile, ProcessMap
    risks = RiskMatrix.objects.filter(
        organization=organization,
        **_build_date_filter('detection_date', date_from, date_to),
    )
    objectives = QualityObjective.objects.filter(
        organization=organization,
        **_build_date_filter('created_at', date_from, date_to),
    )
    documents = Document.objects.filter(
        organization=organization,
        **_build_date_filter('created_at', date_from, date_to),
    )
    stakeholders = StakeholderProfile.objects.filter(
        organization=organization,
        **_build_date_filter('created_at', date_from, date_to),
    )

    return {
        'risks': {
            'total': risks.count(),
            'critical': risks.filter(risk_level='critico').count(),
            'high': risks.filter(risk_level='alto').count(),
            'medium': risks.filter(risk_level='medio').count(),
            'low': risks.filter(risk_level='bajo').count(),
            'open': risks.exclude(status='closed').count(),
            'mitigated': risks.filter(status='mitigated').count(),
        },
        'objectives': {
            'total': objectives.count(),
            'achieved': objectives.filter(status='achieved').count(),
            'in_progress': objectives.filter(status='in_progress').count(),
            'delayed': objectives.filter(status='delayed').count(),
            'active': objectives.filter(status='active').count(),
        },
        'documents': documents.count(),
        'stakeholders': stakeholders.count(),
    }


# ---------------------------------------------------------------------------
# PDF helpers
# ---------------------------------------------------------------------------

_BRAND_BLUE  = (0.0,  0.24, 0.60)   # #003D99
_BRAND_PURPLE = (0.33, 0.18, 0.71)  # #542DB5
_LIGHT_GREY  = (0.95, 0.95, 0.97)
_MID_GREY    = (0.45, 0.45, 0.50)
_WHITE       = (1.0,  1.0,  1.0)
_RED         = (0.78, 0.10, 0.10)
_AMBER       = (0.80, 0.50, 0.05)
_GREEN       = (0.10, 0.55, 0.30)


def _pdf_colors():
    from reportlab.lib import colors
    return {
        'brand_blue':   colors.Color(*_BRAND_BLUE),
        'brand_purple': colors.Color(*_BRAND_PURPLE),
        'light_grey':   colors.Color(*_LIGHT_GREY),
        'mid_grey':     colors.Color(*_MID_GREY),
        'white':        colors.white,
        'red':          colors.Color(*_RED),
        'amber':        colors.Color(*_AMBER),
        'green':        colors.Color(*_GREEN),
        'black':        colors.black,
    }


def _risk_level_color(level, c):
    return {
        'critico': c['red'],
        'alto':    c['amber'],
        'medio':   c['amber'],
        'bajo':    c['green'],
    }.get(level, c['mid_grey'])


def _pdf_cover(elements, styles, org, report_title, user_name, c):
    from reportlab.platypus import Paragraph, Spacer, HRFlowable
    from reportlab.lib.units import cm

    elements.append(Spacer(1, 1.5 * cm))
    elements.append(
        Paragraph(f"<font color='#{_hex(*_BRAND_BLUE)}'>ISOSmart</font>", styles['Title'])
    )
    elements.append(Spacer(1, 0.4 * cm))
    elements.append(Paragraph(report_title, styles['Heading1']))
    elements.append(HRFlowable(width='100%', thickness=2, color=c['brand_blue']))
    elements.append(Spacer(1, 0.6 * cm))
    elements.append(Paragraph(f"<b>Organización:</b> {org.name}", styles['Normal']))
    elements.append(Spacer(1, 0.2 * cm))
    elements.append(Paragraph(f"<b>Generado:</b> {_now_str()}", styles['Normal']))
    elements.append(Spacer(1, 0.2 * cm))
    elements.append(Paragraph(f"<b>Usuario:</b> {user_name}", styles['Normal']))
    elements.append(Spacer(1, 1.0 * cm))


def _hex(r, g, b):
    return '{:02X}{:02X}{:02X}'.format(int(r * 255), int(g * 255), int(b * 255))


def _now_str():
    return timezone.now().strftime('%d/%m/%Y %H:%M UTC')


def _pdf_table_style(c, header_bg=None):
    from reportlab.platypus import TableStyle
    from reportlab.lib import colors
    hbg = header_bg or c['brand_blue']
    return TableStyle([
        ('BACKGROUND',   (0, 0), (-1, 0),  hbg),
        ('TEXTCOLOR',    (0, 0), (-1, 0),  c['white']),
        ('FONTNAME',     (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',     (0, 0), (-1, 0),  9),
        ('ALIGN',        (0, 0), (-1, 0),  'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [c['white'], c['light_grey']]),
        ('FONTSIZE',     (0, 1), (-1, -1), 8),
        ('GRID',         (0, 0), (-1, -1), 0.25, colors.Color(0.8, 0.8, 0.85)),
        ('TOPPADDING',   (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 4),
        ('LEFTPADDING',  (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ])


# ===========================================================================
# PDF generators
# ===========================================================================

def _generate_sgq_executive_pdf(org, stats, filters_desc, user_name):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, HRFlowable

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    c = _pdf_colors()
    elements = []

    _pdf_cover(elements, styles, org, REPORT_LABELS['sgq_executive'], user_name, c)

    if filters_desc:
        elements.append(Paragraph(f"<i>{filters_desc}</i>", styles['Italic']))
        elements.append(Spacer(1, 0.5 * cm))

    # Summary stats table
    elements.append(Paragraph("Resumen de Estado", styles['Heading2']))
    elements.append(Spacer(1, 0.3 * cm))

    summary_data = [
        ['Indicador', 'Total', 'Detalle'],
        ['Riesgos registrados', stats['risks']['total'],
         f"Críticos: {stats['risks']['critical']}  Altos: {stats['risks']['high']}  "
         f"Medios: {stats['risks']['medium']}  Bajos: {stats['risks']['low']}"],
        ['Riesgos abiertos', stats['risks']['open'],
         f"Mitigados: {stats['risks']['mitigated']}"],
        ['Objetivos de calidad', stats['objectives']['total'],
         f"Logrados: {stats['objectives']['achieved']}  En curso: {stats['objectives']['in_progress']}  "
         f"Demorados: {stats['objectives']['delayed']}"],
        ['Documentos', stats['documents'], '—'],
        ['Partes interesadas', stats['stakeholders'], '—'],
    ]

    col_widths = [5.5 * cm, 2.5 * cm, 9 * cm]
    t = Table(summary_data, colWidths=col_widths)
    t.setStyle(_pdf_table_style(c))
    elements.append(t)
    elements.append(Spacer(1, 1.0 * cm))

    # Risk breakdown sub-table
    elements.append(Paragraph("Distribución de Riesgos por Nivel", styles['Heading2']))
    elements.append(Spacer(1, 0.3 * cm))
    risk_data = [
        ['Nivel', 'Cantidad', '% del Total'],
    ]
    total_r = max(stats['risks']['total'], 1)
    for level, key in [('Crítico', 'critical'), ('Alto', 'high'),
                       ('Medio', 'medium'), ('Bajo', 'low')]:
        cnt = stats['risks'][key]
        risk_data.append([level, cnt, f"{100*cnt//total_r}%"])
    t2 = Table(risk_data, colWidths=[5 * cm, 3 * cm, 3 * cm])
    t2.setStyle(_pdf_table_style(c))
    elements.append(t2)
    elements.append(Spacer(1, 1.0 * cm))

    # Objectives breakdown sub-table
    elements.append(Paragraph("Estado de Objetivos de Calidad", styles['Heading2']))
    elements.append(Spacer(1, 0.3 * cm))
    total_o = max(stats['objectives']['total'], 1)
    obj_data = [
        ['Estado', 'Cantidad', '% del Total'],
    ]
    for label, key in [('Logrado', 'achieved'), ('En Progreso', 'in_progress'),
                       ('Demorado', 'delayed'), ('Activo', 'active')]:
        cnt = stats['objectives'][key]
        obj_data.append([label, cnt, f"{100*cnt//total_o}%"])
    t3 = Table(obj_data, colWidths=[5 * cm, 3 * cm, 3 * cm])
    t3.setStyle(_pdf_table_style(c))
    elements.append(t3)

    doc.build(elements)
    buf.seek(0)
    return buf


def _generate_risks_pdf(org, risks, filters_desc, user_name):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    c = _pdf_colors()
    elements = []

    _pdf_cover(elements, styles, org, REPORT_LABELS['risks'], user_name, c)

    if filters_desc:
        elements.append(Paragraph(f"<i>{filters_desc}</i>", styles['Italic']))
        elements.append(Spacer(1, 0.5 * cm))

    elements.append(Paragraph(f"Total de registros: {len(risks)}", styles['Normal']))
    elements.append(Spacer(1, 0.4 * cm))

    headers = ['#', 'Descripción', 'Categoría', 'Probabilidad', 'Impacto',
               'Nivel', 'Estado', 'Responsable', 'Vencimiento']
    data = [headers]

    STATUS_MAP = {
        'identified':     'Identificado',
        'under_analysis': 'En Análisis',
        'mitigated':      'Mitigado',
        'accepted':       'Aceptado',
        'closed':         'Cerrado',
    }

    for i, r in enumerate(risks, 1):
        deadline = r.deadline.strftime('%d/%m/%Y') if r.deadline else '—'
        data.append([
            str(i),
            (r.risk_description or '')[:80],
            r.risk_category or '—',
            r.probability or '—',
            r.impact or '—',
            (r.risk_level or '—').upper(),
            STATUS_MAP.get(r.status, r.status or '—'),
            (r.responsible or '—')[:30],
            deadline,
        ])

    col_widths = [0.6*cm, 6*cm, 3*cm, 2.5*cm, 2.5*cm, 2*cm, 2.5*cm, 3.5*cm, 2.5*cm]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    ts = _pdf_table_style(c)

    # Colour risk level cells
    from reportlab.platypus import TableStyle as TS
    for row_idx, risk in enumerate(risks, 1):
        lvl_color = _risk_level_color(risk.risk_level, c)
        ts.add('BACKGROUND', (5, row_idx), (5, row_idx), lvl_color)
        ts.add('TEXTCOLOR',  (5, row_idx), (5, row_idx), c['white'])
        ts.add('FONTNAME',   (5, row_idx), (5, row_idx), 'Helvetica-Bold')

    t.setStyle(ts)
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return buf


def _generate_objectives_pdf(org, objectives, filters_desc, user_name):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    c = _pdf_colors()
    elements = []

    _pdf_cover(elements, styles, org, REPORT_LABELS['objectives'], user_name, c)

    if filters_desc:
        elements.append(Paragraph(f"<i>{filters_desc}</i>", styles['Italic']))
        elements.append(Spacer(1, 0.5 * cm))

    elements.append(Paragraph(f"Total de registros: {len(objectives)}", styles['Normal']))
    elements.append(Spacer(1, 0.4 * cm))

    STATUS_LABEL = {
        'active':      'Activo',
        'in_progress': 'En Progreso',
        'achieved':    'Logrado',
        'delayed':     'Demorado',
        'cancelled':   'Cancelado',
    }

    headers = ['#', 'Descripción', 'Indicador', 'Línea Base', 'Meta',
               'Actual', 'Avance %', 'Estado', 'Responsable', 'Vencimiento']
    data = [headers]

    for i, o in enumerate(objectives, 1):
        deadline = o.deadline.strftime('%d/%m/%Y') if o.deadline else '—'
        progress = f"{o.progress_percentage:.0f}%" if o.progress_percentage is not None else '—'
        data.append([
            str(i),
            (o.objective_description or '')[:70],
            (o.indicator_name or '—')[:30],
            str(o.baseline_value) if o.baseline_value is not None else '—',
            str(o.target_value) if o.target_value is not None else '—',
            str(o.current_value) if o.current_value is not None else '—',
            progress,
            STATUS_LABEL.get(o.status, o.status or '—'),
            (o.responsible or '—')[:30],
            deadline,
        ])

    col_widths = [0.6*cm, 5.5*cm, 3*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2.5*cm, 3*cm, 2.5*cm]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(_pdf_table_style(c))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return buf


# ===========================================================================
# XLSX generators
# ===========================================================================

def _xlsx_header_style():
    from openpyxl.styles import Font, PatternFill, Alignment
    return {
        'font':      Font(bold=True, color='FFFFFF', size=10),
        'fill':      PatternFill(fill_type='solid', fgColor='003D99'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
    }


def _apply_header(ws, row, headers):
    hs = _xlsx_header_style()
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = hs['font']
        cell.fill = hs['fill']
        cell.alignment = hs['alignment']


def _generate_sgq_executive_xlsx(org, stats, filters_desc, user_name):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()

    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = "Resumen Ejecutivo"

    ws['A1'] = 'ISOSmart — Estado SGQ Ejecutivo'
    ws['A1'].font = Font(bold=True, size=14, color='003D99')
    ws['A2'] = f'Organización: {org.name}'
    ws['A3'] = f'Generado: {_now_str()}'
    ws['A4'] = f'Usuario: {user_name}'
    if filters_desc:
        ws['A5'] = f'Filtros: {filters_desc}'

    ws.append([])
    _apply_header(ws, ws.max_row + 1, ['Indicador', 'Total', 'Detalle'])
    ws.append(['Riesgos registrados', stats['risks']['total'],
               f"Críticos:{stats['risks']['critical']} Altos:{stats['risks']['high']} "
               f"Medios:{stats['risks']['medium']} Bajos:{stats['risks']['low']}"])
    ws.append(['Riesgos abiertos', stats['risks']['open'],
               f"Mitigados:{stats['risks']['mitigated']}"])
    ws.append(['Objetivos de calidad', stats['objectives']['total'],
               f"Logrados:{stats['objectives']['achieved']} En curso:{stats['objectives']['in_progress']} "
               f"Demorados:{stats['objectives']['delayed']}"])
    ws.append(['Documentos', stats['documents'], '—'])
    ws.append(['Partes interesadas', stats['stakeholders'], '—'])
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _generate_risks_xlsx(org, risks, filters_desc, user_name):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Riesgos"

    ws['A1'] = 'ISOSmart — Riesgos y Oportunidades'
    ws['A1'].font = Font(bold=True, size=14, color='003D99')
    ws['A2'] = f'Organización: {org.name}'
    ws['A3'] = f'Generado: {_now_str()}'
    ws['A4'] = f'Usuario: {user_name}'
    if filters_desc:
        ws['A5'] = f'Filtros: {filters_desc}'
    ws.append([])

    headers = ['ID', 'Descripción', 'Categoría', 'Módulo Origen',
               'Probabilidad', 'Impacto', 'Nivel de Riesgo', 'Estado',
               'Responsable', 'Cláusula ISO', 'Fecha Detección', 'Vencimiento',
               'Acciones de Mitigación']
    _apply_header(ws, ws.max_row + 1, headers)

    STATUS_MAP = {
        'identified': 'Identificado', 'under_analysis': 'En Análisis',
        'mitigated': 'Mitigado', 'accepted': 'Aceptado', 'closed': 'Cerrado',
    }
    for r in risks:
        ws.append([
            r.id,
            r.risk_description or '',
            r.risk_category or '',
            r.source_module or '',
            r.probability or '',
            r.impact or '',
            (r.risk_level or '').upper(),
            STATUS_MAP.get(r.status, r.status or ''),
            r.responsible or '',
            r.iso_clause or '',
            r.detection_date.strftime('%d/%m/%Y') if r.detection_date else '',
            r.deadline.strftime('%d/%m/%Y') if r.deadline else '',
            r.mitigation_actions or '',
        ])

    for col, width in zip('ABCDEFGHIJKLM', [8, 50, 20, 15, 14, 14, 16, 14, 20, 14, 16, 14, 40]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _generate_objectives_xlsx(org, objectives, filters_desc, user_name):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Objetivos"

    ws['A1'] = 'ISOSmart — Objetivos y Desempeño'
    ws['A1'].font = Font(bold=True, size=14, color='003D99')
    ws['A2'] = f'Organización: {org.name}'
    ws['A3'] = f'Generado: {_now_str()}'
    ws['A4'] = f'Usuario: {user_name}'
    if filters_desc:
        ws['A5'] = f'Filtros: {filters_desc}'
    ws.append([])

    headers = ['ID', 'Descripción', 'Indicador', 'Unidad Medida',
               'Línea Base', 'Meta', 'Valor Actual', 'Avance %',
               'Frecuencia Medición', 'Estado', 'Responsable',
               'Cláusula ISO', 'Vencimiento', 'Módulo Origen']
    _apply_header(ws, ws.max_row + 1, headers)

    STATUS_LABEL = {
        'active': 'Activo', 'in_progress': 'En Progreso',
        'achieved': 'Logrado', 'delayed': 'Demorado', 'cancelled': 'Cancelado',
    }
    for o in objectives:
        ws.append([
            o.id,
            o.objective_description or '',
            o.indicator_name or '',
            o.measurement_unit or '',
            o.baseline_value,
            o.target_value,
            o.current_value,
            round(o.progress_percentage, 1) if o.progress_percentage is not None else None,
            o.measurement_frequency or '',
            STATUS_LABEL.get(o.status, o.status or ''),
            o.responsible or '',
            o.iso_clause or '',
            o.deadline.strftime('%d/%m/%Y') if o.deadline else '',
            o.source_module or '',
        ])

    for col, width in zip('ABCDEFGHIJKLMN', [8, 55, 30, 15, 12, 12, 12, 10, 20, 14, 20, 14, 14, 15]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ===========================================================================
# CSV generators
# ===========================================================================

def _generate_sgq_executive_csv(org, stats, filters_desc, user_name):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['ISOSmart — Estado SGQ Ejecutivo'])
    w.writerow(['Organización', org.name])
    w.writerow(['Generado', _now_str()])
    w.writerow(['Usuario', user_name])
    if filters_desc:
        w.writerow(['Filtros', filters_desc])
    w.writerow([])
    w.writerow(['Indicador', 'Total', 'Detalle'])
    w.writerow(['Riesgos registrados', stats['risks']['total'],
                f"Criticos:{stats['risks']['critical']} Altos:{stats['risks']['high']} "
                f"Medios:{stats['risks']['medium']} Bajos:{stats['risks']['low']}"])
    w.writerow(['Riesgos abiertos', stats['risks']['open'],
                f"Mitigados:{stats['risks']['mitigated']}"])
    w.writerow(['Objetivos de calidad', stats['objectives']['total'],
                f"Logrados:{stats['objectives']['achieved']} En curso:{stats['objectives']['in_progress']} "
                f"Demorados:{stats['objectives']['delayed']}"])
    w.writerow(['Documentos', stats['documents'], ''])
    w.writerow(['Partes interesadas', stats['stakeholders'], ''])
    return io.BytesIO(buf.getvalue().encode('utf-8-sig'))


def _generate_risks_csv(org, risks, filters_desc, user_name):
    STATUS_MAP = {
        'identified': 'Identificado', 'under_analysis': 'En Analisis',
        'mitigated': 'Mitigado', 'accepted': 'Aceptado', 'closed': 'Cerrado',
    }
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['ISOSmart — Riesgos y Oportunidades'])
    w.writerow(['Organizacion', org.name])
    w.writerow(['Generado', _now_str()])
    w.writerow(['Usuario', user_name])
    if filters_desc:
        w.writerow(['Filtros', filters_desc])
    w.writerow([])
    w.writerow(['ID', 'Descripcion', 'Categoria', 'Modulo Origen',
                'Probabilidad', 'Impacto', 'Nivel Riesgo', 'Estado',
                'Responsable', 'Clausula ISO', 'Fecha Deteccion', 'Vencimiento'])
    for r in risks:
        w.writerow([
            r.id, r.risk_description or '', r.risk_category or '',
            r.source_module or '', r.probability or '', r.impact or '',
            (r.risk_level or '').upper(),
            STATUS_MAP.get(r.status, r.status or ''),
            r.responsible or '', r.iso_clause or '',
            r.detection_date.strftime('%d/%m/%Y') if r.detection_date else '',
            r.deadline.strftime('%d/%m/%Y') if r.deadline else '',
        ])
    return io.BytesIO(buf.getvalue().encode('utf-8-sig'))


def _generate_objectives_csv(org, objectives, filters_desc, user_name):
    STATUS_LABEL = {
        'active': 'Activo', 'in_progress': 'En Progreso',
        'achieved': 'Logrado', 'delayed': 'Demorado', 'cancelled': 'Cancelado',
    }
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['ISOSmart — Objetivos y Desempeno'])
    w.writerow(['Organizacion', org.name])
    w.writerow(['Generado', _now_str()])
    w.writerow(['Usuario', user_name])
    if filters_desc:
        w.writerow(['Filtros', filters_desc])
    w.writerow([])
    w.writerow(['ID', 'Descripcion', 'Indicador', 'Unidad', 'Linea Base',
                'Meta', 'Actual', 'Avance %', 'Estado', 'Responsable',
                'Clausula ISO', 'Vencimiento'])
    for o in objectives:
        progress = round(o.progress_percentage, 1) if o.progress_percentage is not None else ''
        w.writerow([
            o.id, o.objective_description or '', o.indicator_name or '',
            o.measurement_unit or '',
            o.baseline_value if o.baseline_value is not None else '',
            o.target_value if o.target_value is not None else '',
            o.current_value if o.current_value is not None else '',
            progress,
            STATUS_LABEL.get(o.status, o.status or ''),
            o.responsible or '', o.iso_clause or '',
            o.deadline.strftime('%d/%m/%Y') if o.deadline else '',
        ])
    return io.BytesIO(buf.getvalue().encode('utf-8-sig'))


# ===========================================================================
# Public API
# ===========================================================================

CONTENT_TYPES = {
    'pdf':  'application/pdf',
    'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'csv':  'text/csv; charset=utf-8',
}

EXTENSIONS = {'pdf': 'pdf', 'xlsx': 'xlsx', 'csv': 'csv'}


def _filters_description(date_from, date_to, status_filter):
    parts = []
    if date_from:
        parts.append(f"Desde: {date_from}")
    if date_to:
        parts.append(f"Hasta: {date_to}")
    if status_filter:
        parts.append(f"Estado: {status_filter}")
    return "  |  ".join(parts) if parts else ""


def generate_report(organization, report_type, fmt, date_from, date_to,
                    status_filter, user_name):
    """
    Returns (buffer: BytesIO, content_type: str, filename: str).

    Parameters
    ----------
    organization : Organization model instance
    report_type  : 'sgq_executive' | 'risks' | 'objectives'
    fmt          : 'pdf' | 'xlsx' | 'csv'
    date_from    : str 'YYYY-MM-DD' or None
    date_to      : str 'YYYY-MM-DD' or None
    status_filter: str or None
    user_name    : str
    """
    if fmt not in CONTENT_TYPES:
        raise ValueError(f"Formato inválido: {fmt}")
    if report_type not in REPORT_LABELS:
        raise ValueError(f"Tipo de reporte inválido: {report_type}")

    filters_desc = _filters_description(date_from, date_to, status_filter)
    ts = timezone.now().strftime('%Y%m%d_%H%M')
    filename = f"isosmart_{report_type}_{ts}.{EXTENSIONS[fmt]}"
    content_type = CONTENT_TYPES[fmt]

    if report_type == 'sgq_executive':
        stats = _executive_stats(organization, date_from, date_to)
        generators = {
            'pdf':  lambda: _generate_sgq_executive_pdf(organization, stats, filters_desc, user_name),
            'xlsx': lambda: _generate_sgq_executive_xlsx(organization, stats, filters_desc, user_name),
            'csv':  lambda: _generate_sgq_executive_csv(organization, stats, filters_desc, user_name),
        }
    elif report_type == 'risks':
        risks = _risk_data(organization, date_from, date_to, status_filter)
        generators = {
            'pdf':  lambda: _generate_risks_pdf(organization, risks, filters_desc, user_name),
            'xlsx': lambda: _generate_risks_xlsx(organization, risks, filters_desc, user_name),
            'csv':  lambda: _generate_risks_csv(organization, risks, filters_desc, user_name),
        }
    elif report_type == 'objectives':
        objectives = _objective_data(organization, date_from, date_to, status_filter)
        generators = {
            'pdf':  lambda: _generate_objectives_pdf(organization, objectives, filters_desc, user_name),
            'xlsx': lambda: _generate_objectives_xlsx(organization, objectives, filters_desc, user_name),
            'csv':  lambda: _generate_objectives_csv(organization, objectives, filters_desc, user_name),
        }

    buf = generators[fmt]()
    return buf, content_type, filename
